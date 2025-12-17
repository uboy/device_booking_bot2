from __future__ import annotations

import csv
import os
from typing import Dict, List

from openpyxl import load_workbook

REQUIRED_COLUMNS = ("SN", "Name", "Type")
OPTIONAL_GROUP_COLUMNS = ("GroupId", "group_id", "GROUP_ID", "groupid", "GROUPID")


def _read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or not set(REQUIRED_COLUMNS).issubset(reader.fieldnames):
            raise ValueError("В файле должны быть колонки SN, Name, Type")
        group_col = next((c for c in reader.fieldnames if c in OPTIONAL_GROUP_COLUMNS), None)
        rows: List[Dict[str, str]] = []
        for row in reader:
            group_raw = (row.get(group_col) if group_col else "") or ""
            rows.append(
                {
                    "SN": (row.get("SN") or "").strip(),
                    "Name": (row.get("Name") or "").strip(),
                    "Type": (row.get("Type") or "").strip(),
                    "GroupId": str(group_raw).strip(),
                }
            )
        return rows


def _read_xlsx(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
    if not set(REQUIRED_COLUMNS).issubset(header):
        raise ValueError("В файле должны быть колонки SN, Name, Type")

    # Сопоставление индексов колонок
    col_index = {name: header.index(name) for name in REQUIRED_COLUMNS}
    group_col = next((name for name in OPTIONAL_GROUP_COLUMNS if name in header), None)
    if group_col:
        col_index["GroupId"] = header.index(group_col)

    rows: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2):
        def _get(name: str) -> str:
            value = row[col_index[name]].value
            return "" if value is None else str(value).strip()

        group_val = ""
        if group_col and "GroupId" in col_index:
            gv = row[col_index["GroupId"]].value
            group_val = "" if gv is None else str(gv).strip()

        rows.append({"SN": _get("SN"), "Name": _get("Name"), "Type": _get("Type"), "GroupId": group_val})
    return rows


def load_devices_from_file(path: str) -> List[Dict[str, str]]:
    """Читает CSV или XLSX с колонками SN, Name, Type и возвращает список словарей."""
    ext = os.path.splitext(path.lower())[1]
    if ext == ".csv":
        return _read_csv(path)
    if ext in (".xlsx", ".xls"):
        return _read_xlsx(path)
    raise ValueError("Поддерживаются только CSV или XLSX")
